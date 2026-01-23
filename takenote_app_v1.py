
#!/usr/bin/env python3
import platform
import os
import sys
import time
import subprocess
import re
import csv
import difflib
import datetime
import glob
import numpy as np
import socket
import openpyxl
# macOS-specific Tkinter import with error handling
try:
    import tkinter as tk
    from tkinter import *
    from tkinter.filedialog import askopenfilename
    from tkinter import ttk
    from tkinter import messagebox
    import tkinter.font as tkFont
   
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill
    print("✅ Tkinter loaded successfully")
except ImportError as e:
    print(f"❌ Tkinter import failed: {e}")
    print("Please install Python with Tkinter support:")
    print("1. Download from https://www.python.org/downloads/")
    print("2. Or use: brew install python-tk")
    sys.exit(1)

# Optional imports with graceful fallback
try:
    import pyfirmata
    PYFIRMATA_AVAILABLE = True
except ImportError:
    print("⚠️  pyfirmata not available - hardware features disabled")
    PYFIRMATA_AVAILABLE = False

try:
    import bluetooth
    from bluetooth import *
    BLUETOOTH_AVAILABLE = True
except ImportError:
    print("⚠️  pybluez not available - Bluetooth features disabled")
    BLUETOOTH_AVAILABLE = False

# Try to import the conversion module
try:
    import Hex2TxtTakenote_ENG1_aug_2021
    CONVERTER_AVAILABLE = True
except ImportError:
    print("⚠️  Hex2TxtTakenote_ENG1_aug_2021 not available - conversion features disabled")
    CONVERTER_AVAILABLE = False



# Arduino board initialization with auto-detection
import serial.tools.list_ports
from pyfirmata import Arduino, util

# ===== GLOBAL VARIABLES =====
board = None
TARGET_NAME = "TakeNote V_3.2"
PORT =1
DEVICE_CONNECTED = False
CONNECTED_DEVICE_ADDR = None
version_number = ""
sock = None  # Bluetooth socket

# Assign your main UI elements
status_label = None
label1 = None
btn_transfer = None

# ===== ARDUINO CONNECTION =====
def find_arduino_port():
    """Improved Arduino / CH340 / USB-Serial auto-detection."""
    ports = serial.tools.list_ports.comports()
    
    print("\n🔍 Available Ports:")
    for p in ports:
        print(f" - {p.device} | {p.description} | {p.hwid}")

    for port in ports:
        desc = (port.description or "").lower()
        hwid = (port.hwid or "").lower()
        dev  = port.device.lower()

        # Direct Arduino detection
        if "arduino" in desc or "arduino" in hwid:
            return port.device
        
        # CH340 / CH9102 Serial USB detection
        if "ch340" in desc or "ch340" in hwid or "wch" in hwid:
            return port.device
        
        # Generic Linux/macOS port names
        if "ttyusb" in dev or "ttyacm" in dev:
            return port.device

        # Windows generic USB
        if "usb" in desc or "usb" in hwid:
            return port.device

    return None


def connect_arduino():
    """Connect to Arduino using pyFirmata with auto baud rate fallback."""
    global board
    port = find_arduino_port()

    if port is None:
        print("❌ Arduino not detected.")
        return False

    print(f"🔌 Found Arduino on: {port}")

    # Try Firmata standard baudrates
    for baud in [57600, 115200]:
        try:
            print(f"⏳ Trying baudrate {baud}...")
            board = Arduino(port, baudrate=baud)
            it = util.Iterator(board)
            it.start()
            print("✅ Arduino connected successfully!")
            return True
        except Exception as e:
            print(f"⚠️ Failed at {baud}: {e}")

    print("❌ Unable to connect to Arduino.")
    return False

if __name__ == "__main__":
    connect_arduino()

# ===== BRAILLE CONVERSION =====

def alplow_brl(rec):
    """Convert lowercase character to braille pattern"""
    alp_lower_brl = {
        "a": np.array([0,0,1,0,0,0,0,0,0], dtype=bool),
        "b": np.array([0,1,1,0,0,0,0,0,0], dtype=bool),
        "c": np.array([0,0,1,1,0,0,0,0,0], dtype=bool),
        "d": np.array([0,0,1,1,1,0,0,0,0], dtype=bool),
        "e": np.array([0,0,1,0,1,0,0,0,0], dtype=bool),
        "f": np.array([0,1,1,1,0,0,0,0,0], dtype=bool),
        "g": np.array([0,1,1,1,1,0,0,0,0], dtype=bool),
        "h": np.array([0,1,1,0,1,0,0,0,0], dtype=bool),
        "i": np.array([0,1,0,1,0,0,0,0,0], dtype=bool),
        "j": np.array([0,1,0,1,1,0,0,0,0], dtype=bool),
        "k": np.array([1,0,1,0,0,0,0,0,0], dtype=bool),
        "l": np.array([1,1,1,0,0,0,0,0,0], dtype=bool),
        "m": np.array([1,0,1,1,0,0,0,0,0], dtype=bool),
        "n": np.array([1,0,1,1,1,0,0,0,0], dtype=bool),
        "o": np.array([1,0,1,0,1,0,0,0,0], dtype=bool),
        "p": np.array([1,1,1,1,0,0,0,0,0], dtype=bool),
        "q": np.array([1,1,1,1,1,0,0,0,0], dtype=bool),
        "r": np.array([1,1,1,0,1,0,0,0,0], dtype=bool),
        "s": np.array([1,1,0,1,0,0,0,0,0], dtype=bool),
        "t": np.array([1,1,0,1,1,0,0,0,0], dtype=bool),
        "u": np.array([1,0,1,0,0,1,0,0,0], dtype=bool),
        "v": np.array([1,1,1,0,0,1,0,0,0], dtype=bool),
        "w": np.array([0,1,0,1,1,1,0,0,0], dtype=bool),
        "x": np.array([1,0,1,1,0,1,0,0,0], dtype=bool),
        "y": np.array([1,0,1,1,1,1,0,0,0], dtype=bool),
        "z": np.array([1,0,1,0,1,1,0,0,0], dtype=bool),
        "*": np.array([1,0,0,0,1,0,0,0,0], dtype=bool),
        "@": np.array([1,0,0,1,1,0,0,0,0], dtype=bool),
        ":": np.array([0,1,0,0,1,0,0,0,0], dtype=bool),
        ",": np.array([0,1,0,0,0,0,0,0,0], dtype=bool),
        "!": np.array([1,1,0,0,1,0,0,0,0], dtype=bool),
        ".": np.array([0,1,0,0,1,1,0,0,0], dtype=bool),
        "?": np.array([1,1,0,0,0,1,0,0,0], dtype=bool),
        ";": np.array([1,1,0,0,0,0,0,0,0], dtype=bool),
        "/": np.array([1,0,0,1,0,0,0,0,0], dtype=bool),
        "non": np.array([0,0,0,0,0,0,0,0,0], dtype=bool),
        " ": np.array([0,0,0,0,0,0,0,1,0], dtype=bool),
        "\n": np.array([0,0,0,0,0,0,0,0,1], dtype=bool),
        '"': np.array([1,0,0,0,1,1,0,0,0], dtype=bool),
        "'": np.array([1,0,0,0,0,0,0,0,0], dtype=bool),
        "(": np.array([1,1,1,0,1,1,0,0,0], dtype=bool),
        ")": np.array([1,1,0,1,1,1,0,0,0], dtype=bool),
        "&": np.array([1,1,1,1,0,1,0,0,0], dtype=bool),
        "%": np.array([0,0,1,1,0,1,0,0,0], dtype=bool),
        "+": np.array([1,0,0,1,0,1,0,0,0], dtype=bool),
        "-": np.array([1,0,0,0,0,1,0,0,0], dtype=bool),
    }
    return alp_lower_brl.get(rec, alp_lower_brl["non"])

def alpup_brl(rec):
    """Convert uppercase character to braille pattern"""
    alp_upper_brl = {
        "A": np.array([0,0,1,0,0,0,0,0,0], dtype=bool),
        "B": np.array([0,1,1,0,0,0,0,0,0], dtype=bool),
        "C": np.array([0,0,1,1,0,0,0,0,0], dtype=bool),
        "D": np.array([0,0,1,1,1,0,0,0,0], dtype=bool),
        "E": np.array([0,0,1,0,1,0,0,0,0], dtype=bool),
        "F": np.array([0,1,1,1,0,0,0,0,0], dtype=bool),
        "G": np.array([0,1,1,1,1,0,0,0,0], dtype=bool),
        "H": np.array([0,1,1,0,1,0,0,0,0], dtype=bool),
        "I": np.array([0,1,0,1,0,0,0,0,0], dtype=bool),
        "J": np.array([0,1,0,1,1,0,0,0,0], dtype=bool),
        "K": np.array([1,0,1,0,0,0,0,0,0], dtype=bool),
        "L": np.array([1,1,1,0,0,0,0,0,0], dtype=bool),
        "M": np.array([1,0,1,1,0,0,0,0,0], dtype=bool),
        "N": np.array([1,0,1,1,1,0,0,0,0], dtype=bool),
        "O": np.array([1,0,1,0,1,0,0,0,0], dtype=bool),
        "P": np.array([1,1,1,1,0,0,0,0,0], dtype=bool),
        "Q": np.array([1,1,1,1,1,0,0,0,0], dtype=bool),
        "R": np.array([1,1,1,0,1,0,0,0,0], dtype=bool),
        "S": np.array([1,1,0,1,0,0,0,0,0], dtype=bool),
        "T": np.array([1,1,0,1,1,0,0,0,0], dtype=bool),
        "U": np.array([1,0,1,0,0,1,0,0,0], dtype=bool),
        "V": np.array([1,1,1,0,0,1,0,0,0], dtype=bool),
        "W": np.array([0,1,0,1,1,1,0,0,0], dtype=bool),
        "X": np.array([1,0,1,1,0,1,0,0,0], dtype=bool),
        "Y": np.array([1,0,1,1,1,1,0,0,0], dtype=bool),
        "Z": np.array([1,0,1,0,1,1,0,0,0], dtype=bool),
    }
    return alp_upper_brl.get(rec, np.array([0,0,0,0,0,0,0,0,0], dtype=bool))

def num_brl(rec):
    """Convert number to braille pattern"""
    number_brl = {
        "1": np.array([0,0,1,0,0,0,0,0,0], dtype=bool),
        "2": np.array([0,1,1,0,0,0,0,0,0], dtype=bool),
        "3": np.array([0,0,1,1,0,0,0,0,0], dtype=bool),
        "4": np.array([0,0,1,1,1,0,0,0,0], dtype=bool),
        "5": np.array([0,0,1,0,1,0,0,0,0], dtype=bool),
        "6": np.array([0,1,1,1,0,0,0,0,0], dtype=bool),
        "7": np.array([0,1,1,1,1,0,0,0,0], dtype=bool),
        "8": np.array([0,1,1,0,1,0,0,0,0], dtype=bool),
        "9": np.array([0,1,0,1,0,0,0,0,0], dtype=bool),
        "0": np.array([0,1,0,1,1,0,0,0,0], dtype=bool),
    }
    return number_brl.get(rec, np.array([0,0,0,0,0,0,0,0,0], dtype=bool))

# Arduino control functions
def file_open():
    """Send file open command to Arduino"""
    if not board:
        return
    board.digital[5].write(0)
    board.digital[6].write(0)
    board.digital[7].write(0)
    board.digital[8].write(0)
    board.digital[9].write(0)
    board.digital[10].write(0)
    board.digital[2].write(1)
    board.digital[3].write(0)
    board.digital[4].write(1)
    time.sleep(0.5)
    board.digital[2].write(0)
    board.digital[4].write(0)
    time.sleep(0.5)

def file_close():
    """Send file close command to Arduino"""
    if not board:
        return
    board.digital[5].write(0)
    board.digital[6].write(0)
    board.digital[7].write(0)
    board.digital[8].write(0)
    board.digital[9].write(0)
    board.digital[10].write(0)
    board.digital[2].write(1)
    board.digital[3].write(1)
    board.digital[4].write(0)
    time.sleep(0.5)
    board.digital[2].write(0)
    board.digital[3].write(0)
    time.sleep(0.5)

def dot_led(outled):
    """Send braille pattern to Arduino LEDs"""
    if not board:
        return
    board.digital[5].write(outled[0])
    board.digital[6].write(outled[1])
    board.digital[7].write(outled[2])
    board.digital[8].write(outled[3])
    board.digital[9].write(outled[4])
    board.digital[10].write(outled[5])
    board.digital[2].write(outled[6])
    board.digital[3].write(outled[7])
    board.digital[4].write(outled[8])
    time.sleep(0.5)
    board.digital[5].write(0)
    board.digital[6].write(0)
    board.digital[7].write(0)
    board.digital[8].write(0)
    board.digital[9].write(0)
    board.digital[10].write(0)
    board.digital[2].write(0)
    board.digital[3].write(0)
    board.digital[4].write(0)
    time.sleep(0.5)

def al_num():
    """Send alphabet to number mode switch"""
    if not board:
        return
    board.digital[5].write(1)
    board.digital[6].write(0)
    board.digital[7].write(0)
    board.digital[8].write(1)
    board.digital[9].write(1)
    board.digital[10].write(1)
    time.sleep(0.5)
    board.digital[5].write(0)
    board.digital[6].write(0)
    board.digital[7].write(0)
    board.digital[8].write(0)
    board.digital[9].write(0)
    board.digital[10].write(0)
    time.sleep(0.5)

def num_al():
    """Send number to alphabet mode switch"""
    if not board:
        return
    board.digital[5].write(0)
    board.digital[6].write(0)
    board.digital[7].write(0)
    board.digital[8].write(0)
    board.digital[9].write(1)
    board.digital[10].write(1)
    time.sleep(0.5)
    board.digital[5].write(0)
    board.digital[6].write(0)
    board.digital[7].write(0)
    board.digital[8].write(0)
    board.digital[9].write(0)
    board.digital[10].write(0)
    time.sleep(0.5)

def capital():
    """Send capital letter indicator"""
    if not board:
        return
    board.digital[5].write(0)
    board.digital[6].write(0)
    board.digital[7].write(0)
    board.digital[8].write(0)
    board.digital[9].write(0)
    board.digital[10].write(1)
    time.sleep(0.5)
    board.digital[5].write(0)
    board.digital[6].write(0)
    board.digital[7].write(0)
    board.digital[8].write(0)
    board.digital[9].write(0)
    board.digital[10].write(0)
    time.sleep(0.5)
# ===== TEST FUNCTION =====
def test_individual(txt_widget, status_label):
    """Test individual text from a Tkinter Text widget."""
    global board
    if board is None:
        messagebox.showwarning("Board Error", "Arduino not connected. Connect first.")
        return
    
    text_content = txt_widget.get("1.0", tk.END)
    count = 0
    num_flog = 0
    
    file_open()
    time.sleep(1)
    
    for char in text_content[:-1]:
        if char.islower():
            if num_flog == 1: num_al()
            num_flog = 0
            outled = alplow_brl(char)
        elif char.isupper():
            if num_flog == 1: num_al()
            num_flog = 0
            capital()
            outled = alpup_brl(char)
        elif char.isdigit():
            if num_flog == 0: al_num()
            num_flog = 1
            outled = num_brl(char)
        else:
            outled = alplow_brl(char)
        
        dot_led(outled)
        count += 1
        if count % 10 == 0:
            txt_widget.update()
    
    file_close()
    status_label.config(text=f"Hardware testing complete: {count} characters")
    print(f"✅ Hardware testing completed: {count} characters")


#---------------------------Bluetooth Transfer Functions------------------
def find_device(target_name, timeout=8):
    """Find the target Bluetooth device"""
    if not BLUETOOTH_AVAILABLE:
        return None
    
    print(f"🔍 Scanning for {target_name}...")
    try:
        devices = bluetooth.discover_devices(lookup_names=True, duration=timeout)
        for addr, name in devices:
            print(f"  • {name} ({addr})")
            if name and target_name in name:
                return addr
        return None
    except Exception as e:
        print(f"Error during device discovery: {e}")
        return None

def parse_and_save(data_stream):
    """Split incoming stream into separate files based on firmware markers."""
    if not data_stream:
        print("⚠️  No data received.")
        return []
    
    data_str = data_stream.decode(errors='ignore')
    print(f"📄 Received data length: {len(data_str)} characters")
    
    files = re.findall(r'(/NOTE\d+\.txt)(.*?)(?:Continue 01111110 \$)', data_str)
    saved_files = []
    
    if not files:
        print("⚠️  No files found in stream.")
        if data_str.strip():
            filename = f"received_data_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            with open(filename, 'w') as f:
                f.write(data_str.strip())
            saved_files.append(filename)
            print(f"✅ Saved raw data as: {filename}")
        return saved_files
    
    for filename, content in files:
        clean_name = filename.strip('/')
        try:
            with open(clean_name, 'w') as f:
                f.write(content.strip())
            abs_path = os.path.abspath(clean_name)
            saved_files.append(clean_name)
            print(f"✅ Saved file: {abs_path}")
        except Exception as e:
            print(f"❌ Error saving {clean_name}: {e}")
    
    if "File Transfer Done 01111110 $" in data_str:
        print("✅ All files transferred successfully.")
    else:
        print("⚠️  File transfer may be incomplete (missing final marker).")
    
    return saved_files

def receive_all(sock):
    """Read until the transfer is complete, return the complete byte stream."""
    sock.settimeout(10.0)
    buffer = bytearray()
    
    try:
        while True:
            chunk = sock.recv(1024)
            if not chunk:
                break
            print(f"⬇️  Received {len(chunk)} bytes")
            buffer.extend(chunk)
            
            if b"File Transfer Done 01111110 $" in buffer:
                print("✅ Transfer complete marker received.")
                break
    except socket.timeout:
        print("⏱️  Socket timed out (no more data).")
    except Exception as e:
        print(f"⚠️  Receive error: {e}")
    finally:
        sock.close()
    
    return buffer

#---------------------------Hardware Testing Functions------------------



#---------------------------Original Functions-------------------------------
def result():
    """Run data analysis and display results"""
    try:
        # Check if analysis script exists
        analysis_script = "Testjig_data_analysis.py"
        if not os.path.exists(analysis_script):
            txt1.delete("1.0", tk.END)
            txt1.insert(tk.END, f"Error: {analysis_script} not found in current directory")
            return
            
        # Run the external script and capture the output
        process_result = subprocess.run(
            [sys.executable, analysis_script],  # Use current Python interpreter
            capture_output=True, text=True, check=True
        )
        
        # Step 1: Remove ANSI escape codes
        cleaned_output = re.sub(r'\x1B[@-_][0-?]*[ -/]*[@-~]', '', process_result.stdout)
        # Step 2: Remove leading newlines
        cleaned_output = cleaned_output.lstrip('\n')
        # Step 3: Filter out progress and epoch lines
        filtered_output = "\n".join(
            line for line in cleaned_output.splitlines()
            if not re.search(r'^\s*\d+/\d+', line)
            and not re.match(r'^Epoch \d+/\d+', line)
        )
        
        txt1.delete("1.0", tk.END)
        txt1.insert(tk.END, filtered_output)
        
        # Display CSV if it exists
        csv_file = 'mistyped_words.csv'
        if os.path.exists(csv_file):
            display_csv(csv_file, txt1)
        else:
            txt1.insert(tk.END, f"\n\nNote: {csv_file} not found")
            
    except subprocess.CalledProcessError as e:
        txt1.delete("1.0", tk.END)
        txt1.insert(tk.END, f"Error running analysis:\n{e.stderr}")
    except Exception as e:
        txt1.delete("1.0", tk.END)
        txt1.insert(tk.END, f"Unexpected error: {str(e)}")

def display_csv(file_path, text_widget):
    """Display CSV file in a formatted table"""
    try:
        with open(file_path, mode='r', encoding='utf-8') as csv_file:
            reader = csv.reader(csv_file)
            rows = list(reader)
            
            if not rows:
                text_widget.insert(tk.END, "\nThe CSV file is empty.\n")
                return
                
            # Calculate column widths
            column_widths = [max(len(str(item)) for item in col) for col in zip(*rows)]
            border_row = "+" + "+".join(["-" * (width + 2) for width in column_widths]) + "+"
            
            # Configure header style
            text_widget.tag_configure("header", foreground="blue", font=("Courier", 12, "bold"))
            
            # Add table
            text_widget.insert(tk.END, "\n" + border_row + "\n")
            
            # Header row
            header = rows[0]
            formatted_header = "| " + " | ".join(f"{str(item).ljust(width)}" for item, width in zip(header, column_widths)) + " |"
            text_widget.insert(tk.END, formatted_header + "\n", "header")
            text_widget.insert(tk.END, border_row + "\n")
            
            # Data rows
            for row in rows[1:]:
                formatted_row = "| " + " | ".join(f"{str(item).ljust(width)}" for item, width in zip(row, column_widths)) + " |"
                text_widget.insert(tk.END, formatted_row + "\n")
                
            text_widget.insert(tk.END, border_row + "\n")
            
    except FileNotFoundError:
        text_widget.insert(tk.END, f"\nError: {file_path} not found.\n")
    except Exception as e:
        text_widget.insert(tk.END, f"\nError reading {file_path}: {e}\n")


# def convert_hex_to_text(hex_content):
#     """Convert hex data to readable text using the conversion module"""
#     if not CONVERTER_AVAILABLE:
#         print("⚠️  Hex converter not available")
#         return hex_content
    
#     try:
#         # Use the conversion module if available
#         converted_text = Hex2TxtTakenote_ENG1_aug_2021.convert_hex_to_text(hex_content)
#         return converted_text
#     except Exception as e:
#         print(f"⚠️  Hex conversion failed: {e}")
#         return hex_content


def convert_hex_to_text(hex_content):
    """Convert hex data to readable text using Hex2TxtTakenote_ENG1_aug_2021"""
    if not CONVERTER_AVAILABLE:
        print("⚠️  Hex converter not available")
        return hex_content
    
    try:
        # Write hex_content to a temporary file
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, mode="w", suffix=".txt") as tmp_in:
            tmp_in.write(hex_content)
            tmp_in_path = tmp_in.name
        
        tmp_out_path = tmp_in_path.replace(".txt", "_out.txt")
        
        # Call the module’s conversion function
        Hex2TxtTakenote_ENG1_aug_2021.englishGrade1Convert(tmp_in_path, tmp_out_path)
        
        # Read back the converted text
        with open(tmp_out_path, "r", encoding="utf-8") as f:
            converted_text = f.read()
        
        return converted_text
    
    except Exception as e:
        print(f"⚠️  Hex conversion failed: {e}")
        return hex_content

# def get():
#     """Load a text file using file dialog"""
#     try:
#         # Use current directory as initial directory
#         initial_dir = os.getcwd()
        
#         filepath = askopenfilename(
#             initialdir=initial_dir,
#             title="Select a text file",
#             filetypes=(("Text Files", "*.txt"), ("All Files", "*.*"))
#         )
        
#         if not filepath:
#             return
            
#         print(f"Selected file path: {filepath}")
        
#         # Clear the main text area
#         txt1.delete("1.0", tk.END)
        
#         # Read and decode hex values
#         with open(filepath, "r", encoding='utf-8') as input_file:
#             hex_text = input_file.read().strip()
            
#             try:
#                 # Convert hex string to bytes, then decode to text
#                 decoded_text = bytes.fromhex(hex_text).decode('utf-8', errors='replace')
#             except ValueError:
#                 # If file contains non-hex characters, just show raw text
#                 decoded_text = hex_text
            
#             txt1.insert(tk.END, decoded_text)
            
#         # Clear user input area on new file load
#         typed1.delete("1.0", tk.END)
        
#         # Show success message
#         status_label.config(text=f"Loaded: {os.path.basename(filepath)}")
        
#     except Exception as e:
#         messagebox.showerror("Error", f"Failed to load file: {str(e)}")
#         status_label.config(text="Error loading file")

def get():
    """Load a text file using file dialog, decode hex if present, and preserve typed text"""
    try:
        # Use current directory as initial directory
        initial_dir = os.getcwd()
        
        filepath = askopenfilename(
            initialdir=initial_dir,
            title="Select a text file",
            filetypes=(("Text Files", "*.txt"), ("All Files", "*.*"))
        )
        
        if not filepath:
            return  # User cancelled file selection
            
        print(f"Selected file path: {filepath}")
        
        # Read file content
        with open(filepath, "r", encoding='utf-8') as input_file:
            hex_text = input_file.read().strip()
            
            try:
                # Try converting hex to text
                decoded_text = bytes.fromhex(hex_text).decode('utf-8', errors='replace')
            except ValueError:
                # Not hex → use raw text
                decoded_text = hex_text
        
        # Insert content into main text area (append if you want, or replace)
        txt1.delete("1.0", tk.END)  # Clear previous file content
        txt1.insert(tk.END, decoded_text)
        
        # ✅ Preserve typed text area; do NOT delete typed1
        # typed1.delete("1.0", tk.END)  # Removed to keep user input
        
        # Update status
        status_label.config(text=f"Loaded: {os.path.basename(filepath)}")
        
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load file: {str(e)}")
        status_label.config(text="Error loading file")

def test_files_all():
    """Test all text files in the input directory"""
    global time_hour, time_minute, time_second
    
    start_time = time.time()
    count_to = 0
    j = 0
    
    # Use current directory for testing
    input_dir = os.getcwd()
    
    # Count text files
    txt_files = [f for f in os.listdir(input_dir) if f.endswith('.txt')]
    count_to = len(txt_files)
    
    print(f"Number of text files found: {count_to}")
    
    if count_to == 0:
        messagebox.showwarning("Warning", "No text files found in current directory")
        return
    
    # Process each file
    for txt_file in txt_files:
        j += 1
        file_path = os.path.join(input_dir, txt_file)
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file1:
                content = file1.read()
                print(f"Processing file {j}: {txt_file}")
                print(content[:100] + "..." if len(content) > 100 else content)
                
            # Clear and load content
            txt1.delete("1.0", tk.END)
            txt1.insert(tk.END, content)
            
            # Test individual characters
            test_individual()
            
            # Add delay between files
            time.sleep(2)  # Reduced from 7 seconds for faster testing
            
        except Exception as e:
            print(f"Error processing {txt_file}: {e}")
    
    # Calculate elapsed time
    end_time = time.time()
    difference = end_time - start_time
    
    hour = int(difference // 3600)
    difference %= 3600
    minute = int(difference // 60)
    second = int(difference % 60)
    
    print(f"Time taken by TestJig process: {hour} hours, {minute} minutes, {second} seconds")
    
    time_hour = hour
    time_minute = minute
    time_second = second
    
    # Show completion message
    status_label.config(text=f"Testing completed in {minute}m {second}s")
    
    # Call result analysis
    result()

# def transfer_files():
#     """Transfer files via Bluetooth and automatically evaluate"""
#     global version_number
    
#     # Always simulate on macOS if Bluetooth is not available
#     if not BLUETOOTH_AVAILABLE or platform.system() == "Darwin":
#         messagebox.showwarning(
#             "Bluetooth Not Available",
#             "Bluetooth is not supported on macOS or required modules are missing. Simulating file transfer."
#         )
#         simulate_transfer()
#         return

#     try:
#         print(f"🔍 Starting Bluetooth transfer for {TARGET_NAME}...")
#         status_label.config(text="Searching for TakeNote device...")
        
#         # Find the device
#         addr = find_device(TARGET_NAME)
#         if not addr:
#             messagebox.showerror("Device Not Found", f"Could not find '{TARGET_NAME}' device!")
#             status_label.config(text="Device not found")
#             return
        
#         version_number = TARGET_NAME
#         label1.config(text=version_number)
        
#         print(f"🔗 Connecting to {TARGET_NAME} at {addr}:{PORT}")
#         status_label.config(text="Connecting to device...")
        
#         # Create Bluetooth socket
#         sock = socket.socket(
#             socket.AF_BLUETOOTH,
#             socket.SOCK_STREAM,
#             socket.BTPROTO_RFCOMM,
#         )
        
#         # Connect to device
#         sock.connect((addr, PORT))
#         print("✅ Connected!")
#         status_label.config(text="Connected! Requesting files...")
        
#         # Send command to request all files
#         cmd = b'0'  # Request all new files
#         sock.send(cmd)
#         print(f"📡 Sent {cmd!r} → requesting all new files")
        
#         # Receive all data
#         status_label.config(text="Receiving files...")
#         data = receive_all(sock)
        
#         if not data:
#             messagebox.showwarning("No Data", "No data received from device.")
#             status_label.config(text="No data received")
#             return
        
#         # Parse and save files
#         status_label.config(text="Parsing and saving files...")
#         saved_files = parse_and_save(data)
        
#         if not saved_files:
#             messagebox.showwarning("No Files", "No files were saved from the transfer.")
#             status_label.config(text="No files saved")
#             return
        
#         # Process the received files
#         status_label.config(text="Processing received files...")
#         process_transferred_files(saved_files)
        
#         # Show completion message
#         messagebox.showinfo("Transfer Complete", f"Successfully transferred and processed {len(saved_files)} files.")
#         status_label.config(text=f"Transfer completed: {len(saved_files)} files")
        
#     except Exception as e:
#         print(f"❌ Bluetooth transfer error: {e}")
#         messagebox.showerror("Transfer Error", f"Bluetooth transfer failed: {e}")
#         status_label.config(text="Transfer failed")

def scan_for_devices():
    """Scan nearby Bluetooth devices and return list of (name, addr)."""
    try:
        print("🔍 Performing Bluetooth scan...")

        result = subprocess.run(
            ["hcitool", "scan"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )

        output = result.stdout.splitlines()
        devices = []

        for line in output:
            if ":" in line:
                # split by ANY whitespace (tabs OR spaces)
                parts = line.strip().split()
                # Example:
                # AA:BB:CC:DD:EE:FF  DeviceName
                if len(parts) >= 2:
                    addr = parts[0]
                    name = " ".join(parts[1:])
                    devices.append((name.strip(), addr.strip()))

        print(f"📡 Found devices: {devices}")
        return devices

    except Exception as e:
        print(f"❌ Scan error: {e}")
        return []


# ============================================================
#          CONNECT DEVICE + POPUP SELECTION WINDOW
# ============================================================
def connect_device():
    """Scan for ALL Bluetooth devices, show list to user, connect selected."""
    global DEVICE_CONNECTED, CONNECTED_DEVICE_ADDR, version_number, btn_transfer

    status_label.config(text="Scanning for Bluetooth devices...")
    print("🔍 Scanning for Bluetooth devices...")

    # Scan Bluetooth devices
    devices = scan_for_devices()

    # -------------------------------
    # SHOW ALL DEVICES (NO FILTER!)
    # -------------------------------
    available = devices

    if not available:
        messagebox.showerror("No Devices Found", "No Bluetooth devices detected!")
        status_label.config(text="No device found")
        return

    # -------------------------------
    # POPUP for selecting device
    # -------------------------------
    popup = tk.Toplevel()
    popup.title("Select Bluetooth Device")

    # === Center the popup ===
    popup.update_idletasks()
    width = 350
    height = 250
    x = (popup.winfo_screenwidth() // 2) - (width // 2)
    y = (popup.winfo_screenheight() // 2) - (height // 2)
    popup.geometry(f"{width}x{height}+{x}+{y}")
    # =========================

    popup.grab_set()

    tk.Label(
        popup,
        text="Available Bluetooth Devices",
        font=("Arial", 12, "bold")
    ).pack(pady=10)

    listbox = tk.Listbox(popup, width=40, height=8)
    listbox.pack()

    # Add device names
    for name, addr in available:
        listbox.insert(tk.END, f"{name} ({addr})")

    # -------------------------------
    # CONNECT BUTTON INSIDE POPUP
    # -------------------------------
    def select_and_connect():
        global DEVICE_CONNECTED, CONNECTED_DEVICE_ADDR, version_number, btn_transfer

        selection = listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a device.")
            return

        index = selection[0]
        name, addr = available[index]

        popup.destroy()

        print(f"🔗 Connecting to {name} ({addr})...")
        status_label.config(text=f"Connecting to {name}...")

        try:
            # Create Bluetooth RFCOMM socket
            sock = socket.socket(
                socket.AF_BLUETOOTH,
                socket.SOCK_STREAM,
                socket.BTPROTO_RFCOMM,
            )
            sock.connect((addr, PORT))

            # Save connection state
            DEVICE_CONNECTED = True
            CONNECTED_DEVICE_ADDR = addr
            version_number = name
            label1.config(text=f"{version_number} Connected")

            print("✅ Device Connected Successfully!")
            status_label.config(text=f"Connected: {name}")

            messagebox.showinfo("Connected", f"{name} connected successfully!")

            # Disable main connect button
            btn_connect.config(
                state="disabled",
                text="CONNECTED",
                bg="#2d4bb8",
                fg="white"
            )

            btn_disconnect.config(state=tk.NORMAL, bg="#dd4141")

        except Exception as e:
            DEVICE_CONNECTED = False
            CONNECTED_DEVICE_ADDR = None
            print(f"❌ Connection error: {e}")
            messagebox.showerror("Connection Failed", f"Could not connect: {e}")
            status_label.config(text="Connection failed")

    tk.Button(popup, text="Connect", command=select_and_connect).pack(pady=10)



def transfer_files():
    """Transfer files via Bluetooth and automatically evaluate"""
    global version_number

    # Always simulate on macOS if Bluetooth is not available
    if not BLUETOOTH_AVAILABLE or platform.system() == "Darwin":
        messagebox.showwarning(
            "Bluetooth Not Available",
            "Bluetooth is not supported on macOS or required modules are missing. Simulating file transfer."
        )
        simulate_transfer()
        return

    try:
        print("🔍 Checking device connection...")
        status_label.config(text="Checking device connection...")

        # ---------------------------------------------------------
        # 1. Verify device is connected (set in connect_device())
        # ---------------------------------------------------------
        if not DEVICE_CONNECTED:
            messagebox.showerror(
                "Device Not Connected",
                "Please connect the TakeNote device before transferring files!"
            )
            status_label.config(text="Device not connected")
            return

        # Device is connected → proceed
        print(f"Device is connected: {version_number}")

        addr = CONNECTED_DEVICE_ADDR
        print(f"Using device address: {addr}")

        # Update label
        label1.config(text=f"{version_number} Connected")
        status_label.config(text="Preparing transfer...")

        print(f"🔗 Connecting to {addr}:{PORT}")

        # ---------------------------------------------------------
        # 2. Create socket and connect
        # ---------------------------------------------------------
        sock = socket.socket(
            socket.AF_BLUETOOTH,
            socket.SOCK_STREAM,
            socket.BTPROTO_RFCOMM
        )

        sock.connect((addr, PORT))

        print("✅ Connected to TakeNote device!")
        status_label.config(text="Connected! Requesting files...")

        # ---------------------------------------------------------
        # 3. Request all new files
        # ---------------------------------------------------------
        cmd = b'0'    # Command to request all files
        sock.send(cmd)

        print(f"📡 Sent command {cmd!r}")
        status_label.config(text="Receiving files...")

        # ---------------------------------------------------------
        # 4. Receive full data from TakeNote
        # ---------------------------------------------------------
        data = receive_all(sock)

        if not data:
            messagebox.showwarning("No Data", "No data received from device.")
            status_label.config(text="No data received")
            return

        # ---------------------------------------------------------
        # 5. Parse and save files
        # ---------------------------------------------------------
        status_label.config(text="Saving files...")
        saved_files = parse_and_save(data)

        if not saved_files:
            messagebox.showwarning("No Files", "No files were saved.")
            status_label.config(text="No files saved")
            return

        # ---------------------------------------------------------
        # 6. Automatically process received files
        # ---------------------------------------------------------
        status_label.config(text="Processing...")
        process_transferred_files(saved_files)

        # ---------------------------------------------------------
        # 7. Completed successfully
        # ---------------------------------------------------------
        messagebox.showinfo(
            "Transfer Complete",
            f"Successfully transferred and processed {len(saved_files)} files."
        )
        status_label.config(text=f"Transfer complete: {len(saved_files)} files")

    except Exception as e:
        print(f"❌ Transfer Error: {e}")
        messagebox.showerror("Transfer Error", f"Bluetooth transfer failed: {e}")
        status_label.config(text="Transfer failed")

def disconnect_device():
    """Disconnect from the currently connected TakeNote device."""
    global DEVICE_CONNECTED, CONNECTED_DEVICE_ADDR, sock

    if not DEVICE_CONNECTED:
        messagebox.showinfo("Not Connected", "No device is currently connected.")
        return

    try:
        # Close Bluetooth socket safely
        if sock:
            sock.close()
    except Exception as e:
        print(f"Error closing socket: {e}")

    # Reset connection state
    DEVICE_CONNECTED = False
    CONNECTED_DEVICE_ADDR = None
    sock = None

    typed1.delete("1.0", tk.END)
    txt1.delete("1.0", tk.END) 

    # Update UI
    status_label.config(text="Device Disconnected")
    label1.config(text="No Device Connected")
    btn_connect.config(text="CONNECT", state=tk.NORMAL, bg="#4166dd")  # re-enable connect button
    btn_disconnect.config(state=tk.DISABLED, bg="#f75058")  # ← THIS LINE disables the disconnect button
    
    messagebox.showinfo("Disconnected", "Device has been disconnected.")
    print("🔌 Device disconnected.")

def simulate_transfer():
    """Simulate file transfer for testing when Bluetooth is not available"""
    print("Simulating file transfer...")
    status_label.config(text="Simulating file transfer...")
    
    # Look for existing NOTE files in current directory for simulation
    note_files = [f for f in os.listdir('.') if f.startswith('NOTE') and f.endswith('.txt')]
    
    if note_files:
        print(f"Found {len(note_files)} NOTE files for simulation")
        status_label.config(text=f"Found {len(note_files)} files for simulation")
        process_transferred_files(note_files)
    else:
        # Create a sample file for demonstration
        sample_content = "This is a simulated transfer file for testing purposes."
        sample_file = "NOTE1_simulated.txt"
        with open(sample_file, 'w') as f:
            f.write(sample_content)
        print(f"Created sample file: {sample_file}")
        process_transferred_files([sample_file])
    
    # Simulate delay
    window.after(2000, lambda: status_label.config(text="Transfer simulation completed"))

# def process_transferred_files(file_list):
#     """Process the transferred files and load them for evaluation"""
#     if not file_list:
#         return
    
#     # Use the most recent file (or first one) as the typed input
#     latest_file = file_list[0]  # You might want to sort by date/time
    
#     try:
#         with open(latest_file, 'r', encoding='utf-8') as f:
#             content = f.read()
        
#         # Check if content appears to be hex data and convert if possible
#         if content.strip().replace(' ', '').replace('\n', '').replace('\r', '').isalnum() and len(content) > 50:
#             # Might be hex data, try to convert
#             converted_content = convert_hex_to_text(content)
#             if converted_content != content:
#                 print(f"✅ Converted hex data to text")
#                 content = converted_content
        
#         # Load the content into the typing area
#         typed1.delete("1.0", tk.END)
#         typed1.insert(tk.END, content)
        
#         print(f"✅ Loaded transferred file content: {latest_file}")
        
#         # If we have reference text loaded, automatically start comparison
#         if txt1.get("1.0", tk.END).strip():
#             window.after(1000, compare_user_input)  # Small delay then auto-compare
#         else:
#             # Show message to load reference text
#             txt1.delete("1.0", tk.END)
#             txt1.insert(tk.END, f"✅ Transferred file loaded: {latest_file}\n\n")
#             txt1.insert(tk.END, "Please use 'GET FILE' to load a reference text for comparison.\n\n")
#             txt1.insert(tk.END, f"Transferred content preview:\n{content[:200]}{'...' if len(content) > 200 else ''}")
        
#     except Exception as e:
#         print(f"❌ Error processing transferred file {latest_file}: {e}")
#         messagebox.showerror("File Processing Error", f"Error processing transferred file: {e}")

def process_transferred_files(file_list):
    """Process the transferred files and load them into the typed area only"""
    if not file_list:
        return
    
    latest_file = max(file_list, key=os.path.getmtime)
    
    try:
        with open(latest_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Always convert hex if converter is available
        content = convert_hex_to_text(content)
        
        # ✅ Load into typed area only
        typed1.delete("1.0", tk.END)
        typed1.insert(tk.END, content)
        
        print(f"✅ Loaded and converted latest file into typed area: {latest_file}")
        status_label.config(text=f"Typed area loaded: {os.path.basename(latest_file)}")
        
    except Exception as e:
        print(f"❌ Error processing {latest_file}: {e}")
        messagebox.showerror("File Processing Error", f"Error processing latest file: {e}")


# def compare_user_input():
#     """Compare the typed text with the reference text"""
#     reference_text = txt1.get("1.0", tk.END).strip()
#     typed_text = typed1.get("1.0", tk.END).strip()
    
#     if not reference_text:
#         messagebox.showwarning("No Reference", "Please load a reference text file first using 'GET FILE'")
#         return
    
#     if not typed_text:
#         messagebox.showwarning("No Input", "Please enter some text in the typing area")
#         return
    
#     # Clear results area
#     result_text.delete("1.0", tk.END)
    
#     # Calculate accuracy metrics
#     total_chars_ref = len(reference_text)
#     total_chars_typed = len(typed_text)
    
#     # Use difflib for detailed comparison
#     matcher = difflib.SequenceMatcher(None, reference_text, typed_text)
#     similarity_ratio = matcher.ratio()
#     accuracy_percentage = similarity_ratio * 100
    
#     # Count character-level differences
#     errors = 0
#     corrections = []
    
#     for tag, i1, i2, j1, j2 in matcher.get_opcodes():
#         if tag == 'replace':
#             errors += max(i2-i1, j2-j1)
#             corrections.append(f"Position {i1}: '{reference_text[i1:i2]}' → '{typed_text[j1:j2]}'")
#         elif tag == 'delete':
#             errors += i2-i1
#             corrections.append(f"Position {i1}: Missing '{reference_text[i1:i2]}'")
#         elif tag == 'insert':
#             errors += j2-j1
#             corrections.append(f"Position {i1}: Extra '{typed_text[j1:j2]}'")
    
#     # Display results
#     result_text.insert(tk.END, "=== TYPING ANALYSIS RESULTS ===\n\n")
#     result_text.insert(tk.END, f"Reference Text Length: {total_chars_ref} characters\n")
#     result_text.insert(tk.END, f"Typed Text Length: {total_chars_typed} characters\n")
#     result_text.insert(tk.END, f"Accuracy: {accuracy_percentage:.2f}%\n")
#     result_text.insert(tk.END, f"Total Errors: {errors}\n\n")
    
#     if corrections:
#         result_text.insert(tk.END, "=== DETAILED CORRECTIONS ===\n")
#         for correction in corrections[:10]:  # Show first 10 corrections
#             result_text.insert(tk.END, f"{correction}\n")
#         if len(corrections) > 10:
#             result_text.insert(tk.END, f"... and {len(corrections) - 10} more corrections\n")
#     else:
#         result_text.insert(tk.END, "🎉 Perfect match! No corrections needed.\n")
    
#     # Update status
#     status_label.config(text=f"Analysis complete: {accuracy_percentage:.1f}% accuracy")
    
#     print(f"Comparison completed: {accuracy_percentage:.2f}% accuracy")





# Example braille mapping for letters
braille_map = {
    "a": "1", "b": "1,2", "c": "1,4", "d": "1,4,5", "e": "1,5",
    "f": "1,2,4", "g": "1,2,4,5", "h": "1,2,5", "i": "2,4", "j": "2,4,5",
    "k": "1,3", "l": "1,2,3", "m": "1,3,4", "n": "1,3,4,5", "o": "1,3,5",
    "p": "1,2,3,4", "q": "1,2,3,4,5", "r": "1,2,3,5", "s": "2,3,4",
    "t": "2,3,4,5", "u": "1,3,6", "v": "1,2,3,6", "w": "2,4,5,6",
    "x": "1,3,4,6", "y": "1,3,4,5,6", "z": "1,3,5,6",

    # Uppercase letters (Capital sign + letter)
    "A": "6,1",    "B": "6,1,2",   "C": "6,1,4",   "D": "6,1,4,5",  "E": "6,1,5",
    "F": "6,1,2,4","G": "6,1,2,4,5","H": "6,1,2,5","I": "6,2,4",   "J": "6,2,4,5",
    "K": "6,1,3",  "L": "6,1,2,3", "M": "6,1,3,4", "N": "6,1,3,4,5","O": "6,1,3,5",
    "P": "6,1,2,3,4","Q": "6,1,2,3,4,5","R": "6,1,2,3,5","S": "6,2,3,4","T": "6,2,3,4,5",
    "U": "6,1,3,6","V": "6,1,2,3,6","W": "6,2,4,5,6","X": "6,1,3,4,6","Y": "6,1,3,4,5,6",
    "Z": "6,1,3,5,6",

    # Numbers (Number sign + letter)
    "0": "3,4,5,6,2,4,5", "1": "3,4,5,6,1",   "2": "3,4,5,6,1,2",  "3": "3,4,5,6,1,4",  
    "4": "3,4,5,6,1,4,5", "5": "3,4,5,6,1,5", "6": "3,4,5,6,1,2,4","7": "3,4,5,6,1,2,4,5",
    "8": "3,4,5,6,1,2,5", "9": "3,4,5,6,2,4",

   # Extended Punctuation & Symbols (single braille cell format)

",": "2",                 # Comma
";": "2,3",               # Semicolon
":": "2,5",               # Colon
".": "2,5,6",             # Period
"!": "2,3,5",             # Exclamation mark
"?": "2,3,6",             # Question mark
"'": "3",                 # Apostrophe
'"': "5",                 # Quotation mark
"-": "3,6",               # Hyphen / Dash
"(": "2,3,5,6",           # Left parenthesis
")": "2,3,5,6",           # Right parenthesis
"/": "3,4",               # Slash
"\\": "1,2,5,6",          # Backslash
"[": "2,3,6",             # Left bracket
"]": "3,5,6",             # Right bracket
"{": "2,3,4,6",           # Left brace
"}": "1,2,5,6",           # Right brace
"<": "1,2,6",             # Less-than
">": "3,4,5",             # Greater-than
"|": "1,2,5",             # Vertical bar
"@": "4,1",               # At sign
"#": "3,4,5,6",           # Number sign

"$": "4,2,3,4",           # Dollar sign (your format)
"%": "4,3,5,6",           # Percent
"&": "4,1,2,3,4,6",       # Ampersand
"*": "5,3,5",             # Asterisk
"+": "5,2,3,5",           # Plus
"=": "5,2,3,5,6",         # Equals
"~": "4,6",               # Tilde
"_": "4,5,6",             # Underscore
"^": "4,6",               # Caret (same as tilde style)

"£": "1,2,3,4,6",         # Pound symbol
"€": "1,2,4,5,6",         # Euro symbol

" ": "0"                  # Space

}

def compare_user_input():
    reference_text = txt1.get("1.0", tk.END).strip()
    typed_text = typed1.get("1.0", tk.END).strip()

    if not reference_text:
        messagebox.showwarning("No Reference", "Please load a reference text file first using 'GET FILE'")
        return
    if not typed_text:
        messagebox.showwarning("No Input", "Please enter some text in the typing area")
        return

    result_text.delete("1.0", tk.END)

    ref_words = reference_text.split()
    typed_words = typed_text.split()

    result_text.tag_configure("error_char", foreground="red", font=("Consolas", 11, "bold"))
    result_text.tag_configure("success", foreground="green", font=("Consolas", 11, "bold"))

    error_summary = {}
    total_errors = 0

    # ---------------------------
    # TABLE 1 → WORD SUMMARY
    # ---------------------------
    table1_rows = []
    for i, ref_word in enumerate(ref_words):

        typed_word = typed_words[i] if i < len(typed_words) else ""

        if not typed_word:
            desc = f"Missing word '{ref_word}'"
            table1_rows.append((ref_word, "—", desc))
            error_summary[desc] = error_summary.get(desc, 0) + 1
            total_errors += 1
            continue

        if ref_word != typed_word:
            matcher = difflib.SequenceMatcher(None, ref_word, typed_word)
            error_descs = []

            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "replace":
                    desc = f"{ref_word[i1:i2]} → {typed_word[j1:j2]}"
                elif tag == "delete":
                    desc = f"Missing '{ref_word[i1:i2]}'"
                elif tag == "insert":
                    desc = f"Extra '{typed_word[j1:j2]}'"
                else:
                    continue

                error_descs.append(desc)
                error_summary[desc] = error_summary.get(desc, 0) + 1
                total_errors += 1

            table1_rows.append((ref_word, typed_word, ", ".join(error_descs)))

    # Accuracy
    accuracy = difflib.SequenceMatcher(None, reference_text, typed_text).ratio() * 100

    # ---------------------------
    # TOP SUMMARY TEXT
    # ---------------------------
    result_text.insert(tk.END, f"TOTAL ERRORS : {total_errors}\n")
    result_text.insert(tk.END, f"ACCURACY     : {accuracy:.2f}%\n\n")

    if total_errors == 0:
        result_text.insert(tk.END, "🎉 Perfect match! No corrections needed.\n")
        status_label.config(text="Analysis complete: 100% accuracy")
        return

    # ---------------------------
    # PRINT SUMMARY TABLE (NO WRAP)
    # ---------------------------
    result_text.insert(tk.END, "=== SUMMARY TABLE ===\n\n")

    separator = "-" * 75 + "\n"
    result_text.insert(tk.END, separator)

    header = "{:<20} | {:<20} | {:<35}\n".format(
        "Original word", "Typed word", "Error description"
    )
    result_text.insert(tk.END, header)
    result_text.insert(tk.END, separator)

    for ref_word, typed_word, desc in table1_rows:
        line = "{:<20} | {:<20} | {:<35}\n".format(ref_word, typed_word, desc)
        result_text.insert(tk.END, line)

    result_text.insert(tk.END, separator + "\n")

    # ---------------------------
    # DETAILED SUMMARY TABLE
    # ---------------------------
    result_text.insert(tk.END, "=== DETAILED SUMMARY ===\n\n")

    separator2 = "-" * 75 + "\n"
    result_text.insert(tk.END, separator2)

    header2 = "{:<17} | {:<10} | {:<10} | {:<10} | {:<10}\n".format(
    "Error", "Correct keys", "Wrong keys", "Missed keys", "frequency"
)

    result_text.insert(tk.END, header2)
    result_text.insert(tk.END, separator2)

    # === BRAILLE KEY DIFFERENCE CALC ===  
    def keyset(x):
        if x == "-" or x == "" or x is None:
            return set()
        return set(x.split(","))

    for desc, count in error_summary.items():

        # Replace error (a → b)
        if "→" in desc:
            a, b = desc.split("→")
            a = a.strip()
            b = b.strip()

            correct_keys = braille_map.get(a, "-")
            wrong_keys = braille_map.get(b, "-")

            correct_set = keyset(correct_keys)
            wrong_set = keyset(wrong_keys)

            missed_set = correct_set - wrong_set
            missed = ",".join(sorted(missed_set)) if missed_set else "-"

        # Missing a character
        elif "Missing" in desc:
            ch = desc.split("'")[1].lower()
            correct_keys = braille_map.get(ch, "-")
            wrong_keys = "-"
            missed = correct_keys

        # Extra character typed
        elif "Extra" in desc:
            ch = desc.split("'")[1].lower()
            correct_keys = "-"
            wrong_keys = braille_map.get(ch, "-")
            missed = "-"

        else:
            correct_keys = wrong_keys = missed = "-"

        row = "{:<17} | {:<12} | {:<10} | {:<11} | {:<5}\n".format(
            desc, correct_keys, wrong_keys, missed, count
        )
        result_text.insert(tk.END, row)

    result_text.insert(tk.END, separator2)

    # status update
    status_label.config(text=f"Analysis complete: {accuracy:.1f}% accuracy")



def clear_all():
    """Clear all text areas"""
    txt1.delete("1.0", tk.END)
    typed1.delete("1.0", tk.END)
    result_text.delete("1.0", tk.END)
    status_label.config(text="All areas cleared")

# def save_results():
#     """Save the analysis results to a file"""
#     results = result_text.get("1.0", tk.END).strip()
#     if not results:
#         messagebox.showwarning("No Results", "No analysis results to save")
#         return
    
#     try:
#         timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
#         filename = f"Analysis_{timestamp}.txt"
        
#         with open(filename, 'w', encoding='utf-8') as f:
#             f.write(results)
        
#         messagebox.showinfo("Saved", f"Results saved to {filename}")
#         status_label.config(text=f"Results saved: {filename}")
        
#     except Exception as e:
#         messagebox.showerror("Save Error", f"Failed to save results: {e}")


def save_results():
    """Save the analysis results into XLSX with proper structured tables"""
    results = result_text.get("1.0", tk.END).strip()
    if not results:
        messagebox.showwarning("No Results", "No analysis results to save")
        return

    try:
        timestamp = datetime.datetime.now().strftime("%d%m%Y_%H%M%S")
        filename = f"Analysis_{timestamp}.xlsx"

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analysis Report"

        row_idx = 1

        # --- Save TOTAL ERRORS and ACCURACY ---
        for line in results.splitlines():
            if line.startswith("TOTAL ERRORS") or line.startswith("ACCURACY"):
                cell = ws.cell(row=row_idx, column=1, value=line)
                cell.font = Font(bold=True)
                row_idx += 1

        row_idx += 2  # blank space

        # --- Summary Table ---
        if "=== SUMMARY TABLE ===" in results:
            ws.cell(row=row_idx, column=1, value="SUMMARY TABLE").font = Font(bold=True, size=12)
            row_idx += 1

            summary_start = results.splitlines().index("=== SUMMARY TABLE ===") + 2  # skip separator line
            summary_end = summary_start
            lines = results.splitlines()
            while summary_end < len(lines) and not lines[summary_end].startswith("==="):
                summary_end += 1

            for i, line in enumerate(lines[summary_start:summary_end]):
                if '|' in line:
                    columns = [col.strip() for col in line.split("|") if col.strip() != ""]
                    for col_idx, val in enumerate(columns, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.alignment = Alignment(horizontal='center')
                        if i == 0:  # header row
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill("solid", fgColor="DDDDDD")
                    row_idx += 1
            row_idx += 2  # blank space

        # --- Detailed Summary Table ---
        if "=== DETAILED SUMMARY ===" in results:
            ws.cell(row=row_idx, column=1, value="DETAILED SUMMARY").font = Font(bold=True, size=12)
            row_idx += 1

            detail_start = results.splitlines().index("=== DETAILED SUMMARY ===") + 2  # skip separator
            detail_end = detail_start
            while detail_end < len(lines) and not lines[detail_end].startswith("==="):
                detail_end += 1

            for i, line in enumerate(lines[detail_start:detail_end]):
                if '|' in line:
                    columns = [col.strip() for col in line.split("|") if col.strip() != ""]
                    for col_idx, val in enumerate(columns, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.alignment = Alignment(horizontal='center')
                        if i == 0:  # header row
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill("solid", fgColor="DDDDDD")
                    row_idx += 1

        # --- Auto column width ---
        for col_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length + 4

        # Save workbook
        wb.save(filename)

        messagebox.showinfo("Saved", f"Results saved to {filename}")
        status_label.config(text=f"Results saved: {filename}")

    except Exception as e:
        messagebox.showerror("Save Error", f"Failed to save results: {e}")

#---------------------------GUI Setup-------------------------------
print("🚀 Initializing TakeNote GUI...")

# Create main window
window = tk.Tk()
window.title("TakeNote - Braille Text Analysis Tool")
window.geometry("1200x800")
window.configure(bg='#f0f0f0')

# Configure grid weights for responsive design
window.grid_rowconfigure(1, weight=1)
window.grid_columnconfigure(0, weight=1)

# Header frame
header_frame = tk.Frame(window, bg='#2c3e50', height=60)
header_frame.grid(row=0, column=0, sticky='ew', padx=5, pady=5)
header_frame.grid_propagate(False)

# Title and version label
title_font = tkFont.Font(family="Arial", size=16, weight="bold")
title_label = tk.Label(header_frame, text="TakeNote Analysis Tool", 
                      font=title_font, fg='white', bg='#2c3e50')
title_label.pack(side=tk.LEFT, padx=20, pady=15)

label1 = tk.Label(header_frame, text="No Device Connected", font=("Arial", 12), 
                  fg='#ecf0f1', bg='#2c3e50')
label1.pack(side=tk.RIGHT, padx=20, pady=15)

# Main content frame
main_frame = tk.Frame(window, bg='#f0f0f0')
main_frame.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
main_frame.grid_rowconfigure(0, weight=1)
main_frame.grid_columnconfigure(0, weight=1)
main_frame.grid_columnconfigure(1, weight=1)

# Left panel - Reference text
left_frame = tk.LabelFrame(main_frame, text="Reference Text", 
                          font=("Arial", 11, "bold"), bg='#f0f0f0')
left_frame.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
left_frame.grid_rowconfigure(1, weight=1)
left_frame.grid_columnconfigure(0, weight=1)

# Reference text buttons
ref_buttons_frame = tk.Frame(left_frame, bg='#f0f0f0')
ref_buttons_frame.grid(row=0, column=0, sticky='ew', padx=5, pady=5)

btn_get = tk.Button(ref_buttons_frame, text="GET FILE", command=get,
                   bg='#3498db', fg='white', font=("Arial", 10, "bold"),
                   relief=tk.RAISED, bd=2)
btn_get.pack(side=tk.LEFT, padx=5)



btn_test_all = tk.Button(ref_buttons_frame, text="TEST TAKENOTE", command=lambda: test_individual(txt1, status_label),

                        bg="#e76354", fg='white', font=("Arial", 10, "bold"),
                        relief=tk.RAISED, bd=2)
btn_test_all.pack(side=tk.RIGHT, padx=5)

btn_clear = tk.Button(ref_buttons_frame, text="CLEAR ALL", command=clear_all,
                     bg='#95a5a6', fg='white', font=("Arial", 10, "bold"),
                     relief=tk.RAISED, bd=2)
btn_clear.pack(side=tk.LEFT, padx=5)

# Reference text area with scrollbar
ref_text_frame = tk.Frame(left_frame)
ref_text_frame.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
ref_text_frame.grid_rowconfigure(0, weight=1)
ref_text_frame.grid_columnconfigure(0, weight=1)

txt1 = tk.Text(ref_text_frame, wrap=tk.WORD, font=("Consolas", 11),
               bg='white', fg='black', insertbackground='red')
txt1.grid(row=0, column=0, sticky='nsew')

scrollbar1 = tk.Scrollbar(ref_text_frame, orient=tk.VERTICAL, command=txt1.yview)
scrollbar1.grid(row=0, column=1, sticky='ns')
txt1.configure(yscrollcommand=scrollbar1.set)

# Right panel - Typed text and results
right_frame = tk.Frame(main_frame, bg='#f0f0f0')
right_frame.grid(row=0, column=1, sticky='nsew', padx=5, pady=5)
right_frame.grid_rowconfigure(0, weight=1)
right_frame.grid_rowconfigure(1, weight=1)
right_frame.grid_columnconfigure(0, weight=1)

# Typed text section
typed_frame = tk.LabelFrame(right_frame, text="Typed Text / Transferred Data", 
                           font=("Arial", 11, "bold"), bg='#f0f0f0')
typed_frame.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
typed_frame.grid_rowconfigure(1, weight=1)
typed_frame.grid_columnconfigure(0, weight=1)

# Typed text buttons
typed_buttons_frame = tk.Frame(typed_frame, bg='#f0f0f0')
typed_buttons_frame.grid(row=0, column=0, sticky='ew', padx=5, pady=5)

btn_connect = tk.Button(typed_buttons_frame, text="CONNECT", command=connect_device,
                        bg="#4166dd", fg='white', font=("Arial", 10, "bold"),
                        relief=tk.RAISED, bd=2)
btn_connect.pack(side=tk.LEFT, padx=5)

btn_transfer = tk.Button(typed_buttons_frame, text="TRANSFER FILES", command=transfer_files,
                        bg='#27ae60', fg='white', font=("Arial", 10, "bold"),
                        relief=tk.RAISED, bd=2)
btn_transfer.pack(side=tk.LEFT, padx=5)

btn_disconnect = tk.Button(typed_buttons_frame, text="DISCONNECT", command=disconnect_device, state=tk.DISABLED,
                       bg="#f75058", fg='white', font=("Arial", 10, "bold"),
                       relief=tk.RAISED, bd=2)
btn_disconnect.pack(side=tk.RIGHT, padx=5)

# Typed text area with scrollbar
typed_text_frame = tk.Frame(typed_frame)
typed_text_frame.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
typed_text_frame.grid_rowconfigure(0, weight=1)
typed_text_frame.grid_columnconfigure(0, weight=1)

typed1 = tk.Text(typed_text_frame, wrap=tk.WORD, font=("Consolas", 11),
                bg='#f8f9fa', fg='black', insertbackground='blue')
typed1.grid(row=0, column=0, sticky='nsew')

scrollbar2 = tk.Scrollbar(typed_text_frame, orient=tk.VERTICAL, command=typed1.yview)
scrollbar2.grid(row=0, column=1, sticky='ns')
typed1.configure(yscrollcommand=scrollbar2.set)

# Results section
results_frame = tk.LabelFrame(right_frame, text="Analysis Results", 
                             font=("Arial", 11, "bold"), bg='#f0f0f0')
results_frame.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
results_frame.grid_rowconfigure(1, weight=1)
results_frame.grid_columnconfigure(0, weight=1)

# Results buttons
results_buttons_frame = tk.Frame(results_frame, bg='#f0f0f0')
results_buttons_frame.grid(row=0, column=0, sticky='ew', padx=5, pady=5)

btn_result = tk.Button(results_buttons_frame, text="RUN ANALYSIS", command=compare_user_input,
                      bg='#9b59b6', fg='white', font=("Arial", 10, "bold"),
                      relief=tk.RAISED, bd=2)
btn_result.pack(side=tk.LEFT, padx=5)

btn_save = tk.Button(results_buttons_frame, text="SAVE RESULTS", command=save_results,
                    bg='#34495e', fg='white', font=("Arial", 10, "bold"),
                    relief=tk.RAISED, bd=2)
btn_save.pack(side=tk.LEFT, padx=5)

# Results text area with scrollbar
results_text_frame = tk.Frame(results_frame)
results_text_frame.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
results_text_frame.grid_rowconfigure(0, weight=1)
results_text_frame.grid_columnconfigure(0, weight=1)

result_text = tk.Text(results_text_frame, wrap=tk.WORD, font=("Consolas", 10),
                     bg='#2c3e50', fg='#ecf0f1', insertbackground='yellow')
result_text.grid(row=0, column=0, sticky='nsew')

scrollbar3 = tk.Scrollbar(results_text_frame, orient=tk.VERTICAL, command=result_text.yview)
scrollbar3.grid(row=0, column=1, sticky='ns')
result_text.configure(yscrollcommand=scrollbar3.set)

# Status bar
status_frame = tk.Frame(window, bg='#34495e', height=30)
status_frame.grid(row=2, column=0, sticky='ew', padx=5, pady=5)
status_frame.grid_propagate(False)

status_label = tk.Label(status_frame, text="Ready - Load a reference file to begin", 
                       font=("Arial", 10), fg='white', bg='#34495e')
status_label.pack(side=tk.LEFT, padx=10, pady=5)

# Initialize with welcome message
welcome_msg = """Welcome to TakeNote - Braille Text Analysis Tool

📝 Instructions:
1. Click 'GET FILE' to load a reference text file
2. Use 'TRANSFER FILES' to receive text from TakeNote device via Bluetooth
3. Or manually type/paste text in the 'Typed Text' area
4. Click 'COMPARE TEXT' to analyze typing accuracy
5. Use 'RUN ANALYSIS' for detailed statistical analysis

🔧 Features:
• Bluetooth file transfer from TakeNote devices
• Real-time text comparison and accuracy measurement
• Braille pattern conversion for hardware testing
• Batch processing of multiple text files
• Detailed error analysis and corrections

Ready to start your analysis!"""

result_text.insert(tk.END, welcome_msg)

# Handle window closing
def on_closing():
    """Handle window closing event"""
    if messagebox.askokcancel("Quit", "Do you want to quit TakeNote?"):
        window.destroy()

window.protocol("WM_DELETE_WINDOW", on_closing)

# Center the window
window.update_idletasks()
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
x = (screen_width - window.winfo_width()) // 2
y = (screen_height - window.winfo_height()) // 2
window.geometry(f"+{x}+{y}")

print("✅ TakeNote GUI initialized successfully!")
print("🖥️  Starting main event loop...")

# Start the GUI event loop
if __name__ == "__main__":
    try:
        window.mainloop()
    except KeyboardInterrupt:
        print("\n⏹️  Application interrupted by user")
    except Exception as e:
        print(f"❌ Application error: {e}")
    finally:
        print("👋 TakeNote application closed")
